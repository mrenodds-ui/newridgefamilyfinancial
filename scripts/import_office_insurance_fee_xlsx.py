#!/usr/bin/env python3
"""Import office Insurance.xlsx + Fee Schedule spreadsheet into HAL knowledge (no PHI)."""

from __future__ import annotations

import json
import re
import sys
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
NR2 = ROOT / "NewRidgeFinancial2"
sys.path.insert(0, str(NR2))

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl required") from exc

INSURANCE_XLSX = Path(r"E:\OneDrive\Desktop\Insurance.xlsx")
FEE_XLSX = Path(r"E:\OneDrive\Desktop\0) Fee Schedule Spreadsheet.xlsx")
OUT_DIR = ROOT / "docs" / "hal_knowledge"
OUT_INS = OUT_DIR / "NEW_RIDGE_OFFICE_INSURANCE_CONTACTS.md"
OUT_FEE = OUT_DIR / "NEW_RIDGE_FEE_SCHEDULES.md"
OUT_JSON = OUT_DIR / "new_ridge_office_insurance_fee.json"
OUT_FEE_RUNTIME = NR2 / "data" / "fee_schedule.json"
LEARNED_PATH = ROOT / "app_data" / "nr2" / "learned_memories.jsonl"

CDT_CODE_RE = re.compile(r"^D\d{4}", re.I)

STAMP = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def clean(val) -> str:
    if val is None:
        return ""
    text = str(val).replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", " / ", text)
    return text.strip()


def parse_insurance_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    out: list[dict] = []
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        company = clean(row[1] if len(row) > 1 else "")
        carrier = clean(row[2] if len(row) > 2 else "")
        if not company and not carrier:
            continue
        out.append(
            {
                "vyneOrEclaimId": clean(row[0] if len(row) > 0 else ""),
                "company": company,
                "carrierSoftDent": carrier,
                "eligibilityPhone": clean(row[4] if len(row) > 4 else ""),
                "eligibilityWebsite": clean(row[5] if len(row) > 5 else ""),
                "claimMailingAddress": clean(row[7] if len(row) > 7 else ""),
                "claimPhone": clean(row[8] if len(row) > 8 else ""),
                "claimFax": clean(row[9] if len(row) > 9 else ""),
                "claimWebsite": clean(row[10] if len(row) > 10 else ""),
                "verificationBestSource": clean(row[12] if len(row) > 12 else ""),
            }
        )
    return out


def parse_fee_index(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Prefer Index sheet
    sheet_name = next((n for n in wb.sheetnames if "index" in n.lower()), wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # find header
    header_idx = None
    for i, row in enumerate(rows):
        vals = [clean(c).lower() for c in (row or [])]
        if any("insurance company" in v for v in vals) and any("fee schedule" in v for v in vals):
            header_idx = i
            break
    if header_idx is None:
        return []
    header = [clean(c) for c in rows[header_idx]]
    # map columns by header text
    def col(name_part: str) -> int | None:
        for i, h in enumerate(header):
            if name_part.lower() in h.lower():
                return i
        return None

    i_co = col("Insurance Company")
    i_pol = col("Insurance Polic")
    i_fee = col("Fee Schedule")
    i_phone = col("Phone")
    i_addr = col("Address")
    out: list[dict] = []
    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        company = clean(row[i_co]) if i_co is not None and i_co < len(row) else ""
        if not company:
            continue
        out.append(
            {
                "insuranceCompany": company,
                "insurancePolicy": clean(row[i_pol]) if i_pol is not None and i_pol < len(row) else "",
                "feeSchedule": clean(row[i_fee]) if i_fee is not None and i_fee < len(row) else "",
                "phone": clean(row[i_phone]) if i_phone is not None and i_phone < len(row) else "",
                "address": clean(row[i_addr]) if i_addr is not None and i_addr < len(row) else "",
            }
        )
    return out


def parse_fee_columns(path: Path) -> list[str]:
    payload = export_fee_schedule_runtime(path)
    return [str(s.get("name") or "") for s in payload.get("schedules") or [] if s.get("name")]


def _slug_schedule(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
    return (slug[:80] or "schedule")


def _resolve_index_fee_to_schedule_id(fee_name: str, company: str, schedule_ids: set[str]) -> str | None:
    """Map Index 'Fee Schedule' / company text onto Data - Main column ids."""
    hay = f"{fee_name} {company}".lower()
    rules = [
        (r"\bdelta\b", "delta_dental"),
        (r"blue cross|bc&bs|bcbs", "bc_bs_based_price"),
        (r"united health", "united_healthcare"),
        (r"\bameritas\b", "ameritas_dental"),
        (r"carrington|careington", "careington_solutions_ppo"),
        (r"\baetna\b", "aetna_dental"),
        (r"\bmetlife\b", "metlife_dental"),
        (r"\bcigna\b", "cigna_dental_health"),
        (r"\bguardian\b", "guardian_direct"),
        (r"\bhumana\b", "humana"),
        (r"\bgeha\b", "geha_connection_dental"),
        (r"concordia.*tricare|\btricare\b", "united_concordia_region_32_tricare"),
        (r"\bconcordia\b", "united_concordia_region_70_parnet"),
        (r"dentemax", "dentemax_dental"),
        (r"dental health alliance|\bdha\b", "dental_health_alliance"),
        (r"unum|colonial", "unum_colonial_life"),
    ]
    for pat, sid in rules:
        if re.search(pat, hay, re.I) and sid in schedule_ids:
            return sid
    return None


def export_fee_schedule_runtime(path: Path) -> dict:
    """Build HAL runtime fee_schedule.json from Data - Main / Adjusted + Index."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    main_name = "Data - Main 2025" if "Data - Main 2025" in wb.sheetnames else wb.sheetnames[1]
    adj_name = "Data - Adjusted 2025" if "Data - Adjusted 2025" in wb.sheetnames else None
    index_name = next((n for n in wb.sheetnames if "index" in n.lower()), None)

    main_rows = list(wb[main_name].iter_rows(values_only=True))
    header_idx = None
    header: list[str] = []
    for i, row in enumerate(main_rows):
        vals = [clean(c) for c in (row or [])]
        if any("Dental Procedure Code" in v for v in vals):
            header_idx = i
            header = vals
            break
    if header_idx is None:
        wb.close()
        return {"version": 1, "schedules": [], "carrierIndex": [], "codes": {}}

    schedules: list[dict] = []
    col_to_id: dict[int, str] = {}
    for j, raw in enumerate(header):
        if not raw or j < 3:
            continue
        name = re.sub(r"\s+", " ", raw)
        if "Practice" in name and "Amount" in name:
            sid = "practice_amount"
            short = "Practice Amount"
            aliases = ["practice amount", "ucr", "office fee", "practice fee", name, short]
        else:
            short = re.sub(r"^20\d{2}\s*", "", name).strip() or name
            sid = _slug_schedule(short)
            aliases = [short, name]
            for tok in re.findall(r"[A-Za-z][A-Za-z&/]+", short):
                if len(tok) >= 3:
                    aliases.append(tok)
        base = sid
        n = 2
        existing = {s["id"] for s in schedules}
        while sid in existing:
            sid = f"{base}_{n}"
            n += 1
        schedules.append(
            {
                "id": sid,
                "name": name,
                "shortName": short,
                "aliases": sorted({a for a in aliases if a}),
                "column": j,
            }
        )
        col_to_id[j] = sid

    codes: dict[str, dict] = {}
    for row in main_rows[header_idx + 1 :]:
        if not row or len(row) < 3:
            continue
        code = clean(row[2]).upper()
        if not CDT_CODE_RE.match(code):
            continue
        code = code.split(".", 1)[0]
        amounts: dict[str, float] = {}
        for j, sid in col_to_id.items():
            if j >= len(row):
                continue
            val = row[j]
            if val is None or val == "":
                continue
            try:
                amounts[sid] = float(val)
            except (TypeError, ValueError):
                continue
        if not amounts:
            continue
        codes[code] = {"category": clean(row[1]), "amounts": amounts}

    if adj_name:
        adj_rows = list(wb[adj_name].iter_rows(values_only=True))
        adj_header = None
        adj_header_idx = None
        for i, row in enumerate(adj_rows):
            vals = [clean(c) for c in (row or [])]
            if any("Dental Procedure Code" in v for v in vals):
                adj_header = vals
                adj_header_idx = i
                break
        adj_col: dict[int, str] = {}
        if adj_header is not None:
            by_short = {str(s["shortName"]).casefold(): s["id"] for s in schedules}
            for j, raw in enumerate(adj_header):
                if not raw or j < 3:
                    continue
                name = re.sub(r"\s+", " ", raw)
                short = re.sub(r"^20\d{2}\s*", "", name).strip() or name
                if "Practice" in name and "Amount" in name:
                    adj_col[j] = "practice_amount"
                elif short.casefold() in by_short:
                    adj_col[j] = by_short[short.casefold()]
            for row in adj_rows[adj_header_idx + 1 :]:
                if not row or len(row) < 3:
                    continue
                code = clean(row[2]).upper()
                if not CDT_CODE_RE.match(code):
                    continue
                code = code.split(".", 1)[0]
                if code not in codes:
                    continue
                adjusted: dict[str, float] = {}
                for j, sid in adj_col.items():
                    if j >= len(row):
                        continue
                    val = row[j]
                    if val is None or val == "":
                        continue
                    try:
                        adjusted[sid] = float(val)
                    except (TypeError, ValueError):
                        continue
                if adjusted:
                    codes[code]["adjustedAmounts"] = adjusted

    schedule_ids = {str(s["id"]) for s in schedules}
    carrier_index: list[dict] = []
    if index_name:
        idx_rows = list(wb[index_name].iter_rows(values_only=True))
        header_i = None
        for i, row in enumerate(idx_rows):
            vals = [clean(c).lower() for c in (row or [])]
            if any("insurance company" in v for v in vals) and any("fee schedule" in v for v in vals):
                header_i = i
                break
        if header_i is not None:
            for row in idx_rows[header_i + 1 :]:
                if not row:
                    continue
                company = clean(row[1] if len(row) > 1 else "")
                if not company:
                    continue
                fee = clean(row[3] if len(row) > 3 else "")
                carrier_index.append(
                    {
                        "insuranceCompany": company,
                        "insurancePolicy": clean(row[2] if len(row) > 2 else ""),
                        "feeSchedule": fee,
                        "phone": clean(row[4] if len(row) > 4 else ""),
                        "resolvedScheduleId": _resolve_index_fee_to_schedule_id(fee, company, schedule_ids),
                    }
                )

    wb.close()
    return {
        "version": 1,
        "importedAt": STAMP,
        "source": str(path),
        "sheet": main_name,
        "adjustedSheet": adj_name,
        "schedules": [{k: v for k, v in s.items() if k != "column"} for s in schedules],
        "carrierIndex": carrier_index,
        "codes": codes,
    }


def write_insurance_md(rows: list[dict]) -> None:
    lines = [
        "# New Ridge office insurance contacts",
        "",
        f"Source: `{INSURANCE_XLSX}`",
        f"Imported: {STAMP}",
        "",
        "Staff working list of carriers with eligibility/claim contacts. SoftDent `Carrier` names align to InsCo where possible.",
        "",
        f"**{len(rows)}** rows.",
        "",
        "| Company | SoftDent carrier | Vyne / E-claim ID | Eligibility phone | Claim phone | Best verification |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for r in rows:
        lines.append(
            "| {company} | {carrier} | {eid} | {ephone} | {cphone} | {best} |".format(
                company=r["company"] or "—",
                carrier=r["carrierSoftDent"] or "—",
                eid=r["vyneOrEclaimId"] or "—",
                ephone=r["eligibilityPhone"] or "—",
                cphone=r["claimPhone"] or "—",
                best=r["verificationBestSource"] or "—",
            )
        )
    lines.extend(["", "## Claim mailing addresses", ""])
    for r in rows:
        if r["claimMailingAddress"]:
            lines.append(f"- **{r['company'] or r['carrierSoftDent']}**: {r['claimMailingAddress']}")
    lines.append("")
    OUT_INS.write_text("\n".join(lines), encoding="utf-8")


def write_fee_md(index_rows: list[dict], schedule_cols: list[str], *, code_count: int = 0) -> None:
    # unique fee schedule names from index
    fee_names = OrderedDict()
    for r in index_rows:
        fs = r.get("feeSchedule") or ""
        if fs:
            fee_names.setdefault(fs, 0)
            fee_names[fs] += 1

    lines = [
        "# New Ridge fee schedules",
        "",
        f"Source: `{FEE_XLSX}`",
        f"Imported: {STAMP}",
        "",
        "Office fee-schedule workbook. HAL runtime lookup uses `NewRidgeFinancial2/data/fee_schedule.json` "
        f"({code_count} CDT codes × schedule columns). Always cite the schedule name with the amount.",
        "",
        "## Fee schedule columns in Data - Main 2025",
        "",
    ]
    for s in schedule_cols:
        lines.append(f"- {s}")
    lines.extend(
        [
            "",
            f"## Insurance → fee schedule map ({len(index_rows)} rows)",
            "",
            "| Insurance company | Policy / product | Fee schedule | Phone |",
            "| --- | --- | --- | --- |",
        ]
    )
    for r in index_rows:
        lines.append(
            "| {co} | {pol} | {fee} | {phone} |".format(
                co=r["insuranceCompany"] or "—",
                pol=r["insurancePolicy"] or "—",
                fee=r["feeSchedule"] or "—",
                phone=r["phone"] or "—",
            )
        )
    lines.extend(["", "## Distinct fee schedule names", ""])
    for name, n in fee_names.items():
        lines.append(f"- {name} ({n} carriers/products)")
    lines.extend(
        [
            "",
            "## HAL usage",
            "",
            "- Ask with a CDT (e.g. D2740) and carrier/schedule name; HAL looks up `fee_schedule.json`.",
            "- For underpayment / CO-45: identify carrier → schedule column → cite allowed amount from the sheet export.",
            "- Do not invent allowed amounts; if the code/schedule is missing, say so.",
            "- Practice amount column is office UCR/practice fee baseline.",
            "",
        ]
    )
    OUT_FEE.write_text("\n".join(lines), encoding="utf-8")


def learned_rows(
    ins_rows: list[dict],
    index_rows: list[dict],
    schedule_cols: list[str],
    *,
    code_count: int = 0,
) -> list[dict]:
    carriers = sorted({r["carrierSoftDent"] for r in ins_rows if r.get("carrierSoftDent")})
    companies = sorted({r["company"] for r in ins_rows if r.get("company")})
    fee_names = sorted({r["feeSchedule"] for r in index_rows if r.get("feeSchedule")})
    # top local
    local_hint = [
        c
        for c in companies
        if re.search(r"Kansas|Delta|Guardian|MetLife|Cigna|Aetna|BCBS|Blue Cross|Principal|Humana|United|GEHA|Ameritas|Careington", c, re.I)
    ][:20]

    base = {
        "created_at": STAMP,
        "last_verified_at": STAMP,
        "confidence": "high",
        "sensitivity_level": "internal_safe",
        "status": "approved",
        "staleness_rule": "verify_monthly",
        "must_not_override": [
            "guardrails",
            "runtime_status",
            "source_availability",
            "external_submission_policy",
        ],
    }

    def L(mid: str, cat: str, text: str, scope: str, notes: str = "") -> dict:
        row = dict(base)
        row.update(
            {
                "id": mid,
                "category": cat,
                "text": text,
                "source": "office Insurance.xlsx + Fee Schedule Spreadsheet.xlsx",
                "scope": scope,
                "notes": notes,
            }
        )
        return row

    schedule_preview = ", ".join(schedule_cols[:12])
    if len(schedule_cols) > 12:
        schedule_preview += f", … ({len(schedule_cols)} total columns)"

    return [
        L(
            "nr2-office-insurance-workbook",
            "insurance_narratives",
            "New Ridge staff insurance contact workbook is Insurance.xlsx (Desktop). It lists company display names, SoftDent "
            f"Carrier names, Vyne/e-claim IDs, eligibility phones/websites, and claim mailing addresses. {len(ins_rows)} rows; "
            f"{len(carriers)} SoftDent carrier labels. Use this for eligibility/claim routing questions — not member benefits.",
            "insurance_narratives",
            notes="Path may move; keep Desktop copy or copy into docs/hal_knowledge.",
        ),
        L(
            "nr2-office-fee-schedule-workbook",
            "insurance_narratives",
            "New Ridge fee schedules live in '0) Fee Schedule Spreadsheet.xlsx' (Desktop). Index maps insurance company → fee "
            f"schedule name; Data - Main 2025 has CDT rows vs schedules including: {schedule_preview}. "
            f"Runtime export has {code_count} CDT codes in NewRidgeFinancial2/data/fee_schedule.json for HAL lookup_fee_schedule. "
            "Always cite the schedule name with the dollar amount; never invent fees.",
            "insurance_narratives",
        ),
        L(
            "nr2-office-fee-schedule-names",
            "insurance_narratives",
            "Distinct fee schedule names used at New Ridge (from fee workbook index): "
            + (", ".join(fee_names[:40]) if fee_names else "(none parsed)")
            + (f" — {len(fee_names)} total." if fee_names else "."),
            "insurance_narratives",
        ),
        L(
            "nr2-office-insurance-local-mix",
            "insurance_narratives",
            "Office insurance workbook highlights local/common carriers including: "
            + (", ".join(local_hint) if local_hint else ", ".join(companies[:15]))
            + ". Cross-check SoftDent InsCo master and Vyne IDs on the card before claim submit.",
            "insurance_narratives",
        ),
        L(
            "nr2-ops-fee-schedule-scaffold",
            "insurance_narratives",
            "For underpayment/CO-45 or 'what is allowed for D#### on carrier X', use HAL fee schedule lookup "
            "(lookup_fee_schedule / fee_schedule.json from Data - Main 2025). Identify the carrier's schedule column, "
            "cite that schedule name and the exported amount. Do not invent allowed amounts from memory.",
            "insurance_narratives",
            notes="Replaces prior fill-in scaffold with live CDT lookup.",
        ),
    ]


def merge_learned(rows: list[dict]) -> tuple[int, int]:
    existing: set[str] = set()
    if LEARNED_PATH.is_file():
        for line in LEARNED_PATH.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(obj, dict) and obj.get("id"):
                existing.add(str(obj["id"]))
    # Allow replace for fee-schedule memories when re-importing
    replace_ids = {
        "nr2-ops-fee-schedule-scaffold",
        "nr2-office-fee-schedule-workbook",
        "nr2-office-fee-schedule-names",
    }
    kept: list[str] = []
    if LEARNED_PATH.is_file():
        for line in LEARNED_PATH.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                kept.append(line)
                continue
            mid = str(obj.get("id") or "")
            if mid in replace_ids:
                continue
            kept.append(line)
    added = 0
    skipped = 0
    LEARNED_PATH.parent.mkdir(parents=True, exist_ok=True)
    with LEARNED_PATH.open("w", encoding="utf-8", newline="\n") as handle:
        for line in kept:
            handle.write(line + "\n")
        for row in rows:
            mid = str(row["id"])
            if mid in existing and mid not in replace_ids:
                skipped += 1
                continue
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
            added += 1
    return added, skipped


def main() -> int:
    if not INSURANCE_XLSX.is_file():
        print("missing", INSURANCE_XLSX)
        return 1
    if not FEE_XLSX.is_file():
        print("missing", FEE_XLSX)
        return 1

    ins_rows = parse_insurance_xlsx(INSURANCE_XLSX)
    index_rows = parse_fee_index(FEE_XLSX)
    fee_runtime = export_fee_schedule_runtime(FEE_XLSX)
    schedule_cols = [str(s.get("name") or "") for s in fee_runtime.get("schedules") or [] if s.get("name")]
    code_count = len(fee_runtime.get("codes") or {})
    print(f"Insurance rows: {len(ins_rows)}")
    print(f"Fee index rows: {len(index_rows)}")
    print(f"Fee schedule columns: {len(schedule_cols)}")
    print(f"CDT codes exported: {code_count}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_FEE_RUNTIME.parent.mkdir(parents=True, exist_ok=True)
    write_insurance_md(ins_rows)
    write_fee_md(index_rows, schedule_cols, code_count=code_count)
    payload = {
        "importedAt": STAMP,
        "sources": {"insurance": str(INSURANCE_XLSX), "feeSchedule": str(FEE_XLSX)},
        "insuranceContacts": ins_rows,
        "feeScheduleIndex": index_rows,
        "feeScheduleColumnsMain2025": schedule_cols,
        "feeScheduleRuntime": {
            "path": str(OUT_FEE_RUNTIME),
            "codeCount": code_count,
            "scheduleCount": len(schedule_cols),
        },
    }
    OUT_JSON.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    OUT_FEE_RUNTIME.write_text(json.dumps(fee_runtime, ensure_ascii=False), encoding="utf-8")

    mem_rows = learned_rows(ins_rows, index_rows, schedule_cols, code_count=code_count)
    added, skipped = merge_learned(mem_rows)
    try:
        from knowledge_memory_store import write_browser_memo_index_js

        target = write_browser_memo_index_js()
        print(f"Browser index: {target}")
    except Exception as exc:  # pragma: no cover
        print(f"Memo index refresh skipped: {exc}")

    print(f"Wrote {OUT_INS}")
    print(f"Wrote {OUT_FEE}")
    print(f"Wrote {OUT_JSON}")
    print(f"Wrote {OUT_FEE_RUNTIME}")
    print(f"Learned memories: +{added} added/replaced, {skipped} skipped")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
